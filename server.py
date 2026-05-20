from __future__ import annotations

import json
import mimetypes
import os
import base64
import hashlib
import hmac
import re
import secrets
import shutil
import subprocess
import tempfile
import time
import uuid
from io import BytesIO
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse


ROOT = Path(__file__).resolve().parent
EXPORT_DIR = ROOT / "generated-pdfs"
CUSTOM_PRODUCTS_FILE = ROOT / "data" / "custom-products.json"
UPLOAD_DIR = ROOT / "assets" / "custom-products"
USERS_FILE = ROOT / "data" / "users.json"
BASE_CATALOG_FILE = ROOT / "data" / "catalog.json"
USER_PRODUCTS_DIR = ROOT / "data" / "user-products"
USER_QUOTES_DIR = ROOT / "data" / "user-quotes"
SESSION_COOKIE = "quote_session"
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "66778899"
LEGACY_DEFAULT_ADMIN_PASSWORDS = ("360304437",)
SESSIONS: dict[str, str] = {}


def find_chrome() -> str | None:
    candidates = [
        os.environ.get("CHROME_PATH"),
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
        "/usr/bin/google-chrome",
        shutil.which("chrome"),
        shutil.which("chromium"),
        shutil.which("chromium-browser"),
        shutil.which("google-chrome"),
        shutil.which("msedge"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    return None


class QuoteHandler(SimpleHTTPRequestHandler):
    server_version = "QuoteServer/1.0"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/login":
            self.login()
            return

        user = self.current_user()
        if not user:
            self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
            return

        if path == "/api/logout":
            self.logout()
            return

        if path == "/api/users":
            self.create_user(user)
            return

        if path == "/api/delete-user":
            self.delete_user(user)
            return

        if path == "/api/delete-quote-record":
            self.delete_quote_record(user)
            return

        if path == "/api/change-password":
            self.change_password(user)
            return

        if path == "/api/custom-products":
            self.save_custom_products()
            return

        if path == "/api/quote-state":
            self.save_quote_state()
            return

        if path == "/api/product-price":
            self.save_product_price()
            return

        if path == "/api/upload-image":
            self.save_uploaded_image()
            return

        if path == "/api/import-products":
            self.import_products()
            return

        if path != "/export-pdf":
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            html = str(payload.get("html", ""))
            width = max(320, min(int(payload.get("width", 1320)), 4000))
            height = max(320, min(int(payload.get("height", 1200)), 8000))
            filename = sanitize_filename(str(payload.get("filename", "报价单.pdf")))
            pdf = render_pdf(html, width, height)
            if payload.get("returnUrl"):
                EXPORT_DIR.mkdir(exist_ok=True)
                cleanup_exports()
                output_name = f"quote-{uuid.uuid4().hex[:12]}.pdf"
                output_path = EXPORT_DIR / output_name
                output_path.write_bytes(pdf)
                body = json.dumps(
                    {"url": f"/generated-pdfs/{output_name}", "filename": filename},
                    ensure_ascii=False,
                ).encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return
        except Exception as exc:
            message = f"PDF export failed: {exc}"
            self.send_response(HTTPStatus.INTERNAL_SERVER_ERROR)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write(message.encode("utf-8"))
            return

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Length", str(len(pdf)))
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote_filename(filename)}")
        self.end_headers()
        self.wfile.write(pdf)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/session":
            user = self.current_user()
            if not user:
                self.send_json({"authenticated": False}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_session(user)
            return

        if self.requires_login(path) and not self.current_user():
            self.redirect("/login.html")
            return

        user = self.current_user()
        if path == "/api/custom-products":
            if not user:
                self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_custom_products()
            return
        if path == "/api/quote-state":
            if not user:
                self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_quote_state()
            return
        if path == "/api/users":
            if not user:
                self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_users(user)
            return
        if path == "/api/quote-records":
            if not user:
                self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_quote_records(user)
            return
        if path == "/api/import-template":
            if not user:
                self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self.send_import_template()
            return
        super().do_GET()

    def requires_login(self, path: str) -> bool:
        public_paths = {
            "/login.html",
            "/login.css",
            "/login.js",
            "/favicon.ico",
        }
        if path in public_paths:
            return False
        if path in ("/", "/index.html", "/manage.html", "/quote-records.html"):
            return True
        if path.startswith("/generated-pdfs/"):
            return True
        return False

    def current_user(self) -> dict | None:
        cookie = self.headers.get("Cookie", "")
        token = ""
        for part in cookie.split(";"):
            name, _, value = part.strip().partition("=")
            if name == SESSION_COOKIE:
                token = value
                break
        username = SESSIONS.get(token)
        if not username:
            return None
        return get_user(username)

    def target_username(self) -> str:
        user = self.current_user()
        if not user:
            raise PermissionError("unauthorized")
        query = parse_qs(urlparse(self.path).query)
        requested = (query.get("user") or [""])[0].strip()
        if user.get("role") == "admin" and requested:
            if not get_user(requested):
                raise ValueError("user not found")
            return requested
        if requested and requested != user["username"]:
            raise PermissionError("forbidden")
        return user["username"]

    def login(self) -> None:
        try:
            payload = self.read_json_body(20_000)
            username = str(payload.get("username", "")).strip()
            password = str(payload.get("password", ""))
            user = get_user(username)
            if not user or not verify_password(password, user.get("passwordHash", "")):
                self.send_json({"error": "账号或密码错误"}, HTTPStatus.UNAUTHORIZED)
                return

            token = secrets.token_urlsafe(32)
            SESSIONS[token] = username
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Set-Cookie", f"{SESSION_COOKIE}={token}; Path=/; HttpOnly; SameSite=Lax")
            body = json.dumps(session_payload(user), ensure_ascii=False).encode("utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"error": f"login failed: {exc}"}, HTTPStatus.BAD_REQUEST)

    def logout(self) -> None:
        cookie = self.headers.get("Cookie", "")
        for part in cookie.split(";"):
            name, _, value = part.strip().partition("=")
            if name == SESSION_COOKIE:
                SESSIONS.pop(value, None)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
        body = b'{"ok":true}'
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_session(self, user: dict) -> None:
        self.send_json(session_payload(user))

    def send_users(self, user: dict) -> None:
        if user.get("role") != "admin":
            self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
            return
        self.send_json({"users": public_users()})

    def send_quote_records(self, user: dict) -> None:
        if user.get("role") != "admin":
            self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
            return
        query = (parse_qs(urlparse(self.path).query).get("q") or [""])[0]
        self.send_json({"records": quote_records(query)})

    def create_user(self, user: dict) -> None:
        if user.get("role") != "admin":
            self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
            return
        try:
            payload = self.read_json_body(50_000)
            username = normalize_username(str(payload.get("username", "")))
            password = str(payload.get("password", ""))
            display_name = str(payload.get("displayName", "")).strip() or username
            if not username:
                raise ValueError("账号不能为空")
            if username == DEFAULT_ADMIN_USERNAME:
                raise ValueError("不能创建同名管理员账号")
            if len(password) < 4:
                raise ValueError("密码至少 4 位")

            data = load_users()
            if username in data["users"]:
                raise ValueError("账号已存在")
            data["users"][username] = {
                "username": username,
                "displayName": display_name,
                "role": "user",
                "passwordHash": hash_password(password),
                "createdAt": int(time.time()),
            }
            save_users(data)
            write_user_products(username, admin_default_products())
            write_user_quote(username, default_quote_state())
            self.send_json({"ok": True, "users": public_users()})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def delete_user(self, user: dict) -> None:
        if user.get("role") != "admin":
            self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
            return
        try:
            payload = self.read_json_body(50_000)
            username = normalize_username(str(payload.get("username", "")))
            if not username:
                raise ValueError("账号不能为空")
            if username == DEFAULT_ADMIN_USERNAME:
                raise ValueError("管理员账号不能删除")
            data = load_users()
            if username not in data["users"]:
                raise ValueError("用户不存在")
            del data["users"][username]
            save_users(data)
            for token, session_username in list(SESSIONS.items()):
                if session_username == username:
                    SESSIONS.pop(token, None)
            remove_user_files(username)
            self.send_json({"ok": True, "users": public_users()})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def delete_quote_record(self, user: dict) -> None:
        if user.get("role") != "admin":
            self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
            return
        try:
            payload = self.read_json_body(50_000)
            username = normalize_username(str(payload.get("username", "")))
            quote_code = str(payload.get("quoteCode", "")).strip()
            if not username and quote_code:
                match = next((record for record in quote_records(quote_code) if record.get("quoteCode") == quote_code), None)
                username = match.get("owner", "") if match else ""
            if not username:
                raise ValueError("报价记录不存在")
            if not get_user(username):
                raise ValueError("用户不存在")
            state = read_user_quote(username)
            if quote_code and state.get("quoteCode") != quote_code:
                raise ValueError("报价记录不存在")

            empty = default_quote_state()
            empty["owner"] = username
            empty["updatedBy"] = user.get("username", DEFAULT_ADMIN_USERNAME)
            empty["updatedAt"] = int(time.time())
            write_user_quote(username, empty)
            self.send_json({"ok": True, "records": quote_records()})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def change_password(self, user: dict) -> None:
        try:
            payload = self.read_json_body(50_000)
            old_password = str(payload.get("oldPassword", ""))
            new_password = str(payload.get("newPassword", ""))
            if not verify_password(old_password, user.get("passwordHash", "")):
                raise ValueError("原密码不正确")
            if len(new_password) < 4:
                raise ValueError("新密码至少 4 位")
            data = load_users()
            username = user["username"]
            if username not in data["users"]:
                raise ValueError("用户不存在")
            data["users"][username]["passwordHash"] = hash_password(new_password)
            save_users(data)
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def read_json_body(self, max_size: int) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if length > max_size:
            raise ValueError("payload is too large")
        payload = json.loads(self.rfile.read(length).decode("utf-8"))
        if not isinstance(payload, dict):
            raise ValueError("payload must be an object")
        return payload

    def send_json(self, payload: object, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def redirect(self, location: str) -> None:
        self.send_response(HTTPStatus.FOUND)
        self.send_header("Location", location)
        self.end_headers()

    def send_custom_products(self) -> None:
        try:
            body = read_user_products(self.target_username()).encode("utf-8")
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)
            return
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def save_custom_products(self) -> None:
        try:
            payload = self.read_json_body(5_000_000)
            data = normalize_custom_products(payload)
            write_user_products(self.target_username(), data)
            body = json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            body = f"Save failed: {exc}".encode("utf-8")
            self.send_response(HTTPStatus.BAD_REQUEST)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    def send_quote_state(self) -> None:
        try:
            data = read_user_quote(self.target_username())
            self.send_json(data)
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.FORBIDDEN)

    def save_quote_state(self) -> None:
        try:
            payload = self.read_json_body(2_000_000)
            target = self.target_username()
            actor = self.current_user() or {}
            data = normalize_quote_state(payload)
            data["owner"] = target
            data["updatedBy"] = actor.get("username", target)
            data["updatedAt"] = int(time.time())
            write_user_quote(target, data)
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def save_product_price(self) -> None:
        try:
            payload = self.read_json_body(50_000)
            product_id = str(payload.get("productId", "")).strip()
            price = normalize_price(payload.get("price"))
            if not product_id:
                raise ValueError("productId is required")
            if price is None or float(price) < 0:
                raise ValueError("price is invalid")
            target = self.target_username()
            product = update_user_product_price(target, product_id, price)
            self.send_json({"ok": True, "product": product})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def save_uploaded_image(self) -> None:
        try:
            payload = self.read_json_body(12_000_000)
            filename = str(payload.get("filename", "product-image"))
            data_url = str(payload.get("dataUrl", ""))
            match = re.match(r"^data:(image/(?:png|jpeg|webp|gif));base64,(.+)$", data_url, re.DOTALL)
            if not match:
                raise ValueError("unsupported image format")

            mime_type, encoded = match.groups()
            image_bytes = base64.b64decode(encoded, validate=True)
            if not image_bytes or len(image_bytes) > 8 * 1024 * 1024:
                raise ValueError("image is too large")

            ext = {
                "image/png": ".png",
                "image/jpeg": ".jpg",
                "image/webp": ".webp",
                "image/gif": ".gif",
            }[mime_type]
            stem = re.sub(r"[^A-Za-z0-9._-]+", "-", Path(filename).stem).strip(".-")[:60]
            if not stem:
                stem = "product-image"
            output_name = f"{int(time.time())}-{uuid.uuid4().hex[:8]}-{stem}{ext}"
            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            output_path = UPLOAD_DIR / output_name
            output_path.write_bytes(image_bytes)

            body = json.dumps(
                {"url": f"assets/custom-products/{output_name}"},
                ensure_ascii=False,
            ).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            body = f"Upload failed: {exc}".encode("utf-8")
            self.send_response(HTTPStatus.BAD_REQUEST)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    def send_import_template(self) -> None:
        try:
            content = build_import_template()
            filename = "产品快速导入模板.xlsx"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Length", str(len(content)))
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote_filename(filename)}")
            self.end_headers()
            self.wfile.write(content)
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def import_products(self) -> None:
        try:
            payload = self.read_json_body(30_000_000)
            result = import_products_from_payload(
                payload,
                self.target_username(),
                read_current_products_data(self.target_username()),
            )
            write_user_products(self.target_username(), result["data"])
            self.send_json(
                {
                    "ok": True,
                    "imported": result["imported"],
                    "created": result["created"],
                    "updated": result["updated"],
                    "skipped": result["skipped"],
                    "errors": result["errors"],
                }
            )
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)


def ensure_users_file() -> dict:
    USERS_FILE.parent.mkdir(exist_ok=True)
    if USERS_FILE.exists():
        try:
            data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict) and isinstance(data.get("users"), dict):
                users = data["users"]
            else:
                users = {}
        except json.JSONDecodeError:
            users = {}
    else:
        users = {}

    changed = False
    admin = users.get(DEFAULT_ADMIN_USERNAME)
    if not admin:
        users[DEFAULT_ADMIN_USERNAME] = {
            "username": DEFAULT_ADMIN_USERNAME,
            "displayName": "管理员",
            "role": "admin",
            "passwordHash": hash_password(DEFAULT_ADMIN_PASSWORD),
            "createdAt": int(time.time()),
        }
        changed = True
    else:
        if admin.get("role") != "admin":
            admin["role"] = "admin"
            changed = True
        password_hash = str(admin.get("passwordHash", ""))
        if not password_hash or any(verify_password(password, password_hash) for password in LEGACY_DEFAULT_ADMIN_PASSWORDS):
            admin["passwordHash"] = hash_password(DEFAULT_ADMIN_PASSWORD)
            changed = True

    data = {"users": users}
    if changed or not USERS_FILE.exists():
        save_users(data)
    ensure_user_data_files(data)
    return data


def load_users() -> dict:
    return ensure_users_file()


def save_users(data: dict) -> None:
    USERS_FILE.parent.mkdir(exist_ok=True)
    USERS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def get_user(username: str) -> dict | None:
    return load_users().get("users", {}).get(username)


def public_users() -> list[dict]:
    users = load_users().get("users", {})
    return [
        {
            "username": user.get("username", username),
            "displayName": user.get("displayName", username),
            "role": user.get("role", "user"),
        }
        for username, user in sorted(users.items(), key=lambda item: (item[1].get("role") != "admin", item[0]))
    ]


def session_payload(user: dict) -> dict:
    payload = {
        "authenticated": True,
        "username": user["username"],
        "displayName": user.get("displayName", user["username"]),
        "role": user.get("role", "user"),
    }
    if payload["role"] == "admin":
        payload["users"] = public_users()
    return payload


def user_display_name(username: str) -> str:
    user = get_user(username)
    return user.get("displayName", username) if user else username


def quote_records(query: str = "") -> list[dict]:
    users = load_users().get("users", {})
    records = []
    for username, user in users.items():
        state = read_user_quote(username)
        if not quote_state_has_content(state):
            continue
        records.append(quote_record_from_state(username, user, state))

    records.sort(key=lambda record: (record.get("updatedAt") or 0, record.get("quoteCode", "")), reverse=True)
    query = query.strip().lower()
    if not query:
        return records

    def matches(record: dict) -> bool:
        searchable = " ".join(
            str(record.get(field, ""))
            for field in (
                "quoteCode",
                "owner",
                "ownerDisplayName",
                "updatedBy",
                "updatedByDisplayName",
                "customerCompany",
                "customerName",
                "customerPhone",
            )
        ).lower()
        return query in searchable

    return [record for record in records if matches(record)]


def quote_record_from_state(username: str, user: dict, state: dict) -> dict:
    customer = state.get("customer") if isinstance(state.get("customer"), dict) else {}
    subtotal = quote_subtotal(state.get("quote", []))
    offer = normalize_price(state.get("offer"))
    updated_by = normalize_username(str(state.get("updatedBy", ""))) or username
    updated_at = normalize_timestamp(state.get("updatedAt")) or quote_file_mtime(username)
    return {
        "quoteCode": str(state.get("quoteCode") or ""),
        "owner": username,
        "ownerDisplayName": user.get("displayName", username),
        "updatedBy": updated_by,
        "updatedByDisplayName": user_display_name(updated_by),
        "updatedAt": updated_at,
        "customerCompany": str(customer.get("customerCompany") or ""),
        "customerName": str(customer.get("customerName") or ""),
        "customerPhone": str(customer.get("customerPhone") or ""),
        "itemCount": len(state.get("quote", [])) if isinstance(state.get("quote"), list) else 0,
        "subtotal": subtotal,
        "offer": offer,
        "total": offer if offer is not None else subtotal,
    }


def quote_state_has_content(state: dict) -> bool:
    customer = state.get("customer") if isinstance(state.get("customer"), dict) else {}
    return bool(
        state.get("quoteCode")
        or state.get("quote")
        or state.get("offer")
        or any(str(value).strip() for value in customer.values())
    )


def quote_subtotal(items: object) -> int | float:
    if not isinstance(items, list):
        return 0
    total = 0.0
    for item in items:
        if not isinstance(item, dict):
            continue
        price = normalize_price(item.get("price")) or 0
        qty = normalize_price(item.get("qty")) or 0
        total += float(price) * float(qty)
    return int(total) if total.is_integer() else round(total, 2)


def normalize_timestamp(value: object) -> int:
    try:
        timestamp = int(float(value))
    except (TypeError, ValueError):
        return 0
    return timestamp if timestamp > 0 else 0


def normalize_username(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "", value.strip())[:32]


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return f"pbkdf2_sha256${salt}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algorithm, salt, digest = stored.split("$", 2)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    check = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000).hex()
    return hmac.compare_digest(check, digest)


def user_file_stem(username: str) -> str:
    clean = normalize_username(username)
    if not clean:
        raise ValueError("invalid username")
    return clean


def default_custom_products() -> dict:
    return {"overrides": {}, "additions": [], "categoryOrder": [], "productOrder": [], "deletedProductIds": []}


def admin_default_products() -> dict:
    try:
        return normalize_custom_products(json.loads(read_user_products(DEFAULT_ADMIN_USERNAME)))
    except Exception:
        return default_custom_products()


def read_user_products(username: str) -> str:
    path = USER_PRODUCTS_DIR / f"{user_file_stem(username)}.json"
    if path.exists():
        return path.read_text(encoding="utf-8")
    if username == DEFAULT_ADMIN_USERNAME and CUSTOM_PRODUCTS_FILE.exists():
        return CUSTOM_PRODUCTS_FILE.read_text(encoding="utf-8")
    return json.dumps(default_custom_products(), ensure_ascii=False)


def write_user_products(username: str, data: dict) -> None:
    USER_PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)
    path = USER_PRODUCTS_DIR / f"{user_file_stem(username)}.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def user_quote_path(username: str) -> Path:
    return USER_QUOTES_DIR / f"{user_file_stem(username)}.json"


def quote_file_mtime(username: str) -> int:
    path = user_quote_path(username)
    try:
        return int(path.stat().st_mtime) if path.exists() else 0
    except OSError:
        return 0


def default_quote_state() -> dict:
    return {"quote": [], "customer": {}, "offer": "", "terms": {}, "quoteCode": ""}


def read_user_quote(username: str) -> dict:
    path = user_quote_path(username)
    if not path.exists():
        return default_quote_state()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return normalize_quote_state(payload)
    except json.JSONDecodeError:
        return default_quote_state()


def write_user_quote(username: str, data: dict) -> None:
    USER_QUOTES_DIR.mkdir(parents=True, exist_ok=True)
    path = user_quote_path(username)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def remove_user_files(username: str) -> None:
    stem = user_file_stem(username)
    for folder in (USER_PRODUCTS_DIR, USER_QUOTES_DIR):
        path = folder / f"{stem}.json"
        try:
            if path.exists():
                path.unlink()
        except OSError:
            pass


def ensure_user_data_files(data: dict) -> None:
    users = data.get("users", {})
    if not isinstance(users, dict):
        return
    USER_PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)
    USER_QUOTES_DIR.mkdir(parents=True, exist_ok=True)
    for username in users:
        stem = user_file_stem(username)
        product_path = USER_PRODUCTS_DIR / f"{stem}.json"
        quote_path = USER_QUOTES_DIR / f"{stem}.json"
        if not product_path.exists():
            if username == DEFAULT_ADMIN_USERNAME and CUSTOM_PRODUCTS_FILE.exists():
                try:
                    payload = json.loads(CUSTOM_PRODUCTS_FILE.read_text(encoding="utf-8"))
                    write_user_products(username, normalize_custom_products(payload))
                except json.JSONDecodeError:
                    write_user_products(username, default_custom_products())
            else:
                write_user_products(username, admin_default_products())
        if not quote_path.exists():
            write_user_quote(username, default_quote_state())


IMPORT_COLUMNS = [
    ("category", "类别"),
    ("series", "系列"),
    ("name", "品名"),
    ("model", "型号"),
    ("price", "挂牌价"),
    ("productImage", "产品图片"),
    ("details", "主要参数"),
    ("features", "功能介绍"),
]

HEADER_ALIASES = {
    "类别": "category",
    "品类": "category",
    "产品类别": "category",
    "系列": "series",
    "品名": "name",
    "名称": "name",
    "产品名称": "name",
    "型号": "model",
    "型 号": "model",
    "产品型号": "model",
    "挂牌价": "price",
    "价格": "price",
    "单价": "price",
    "产品图片": "productImage",
    "图片": "productImage",
    "图片链接": "productImage",
    "主要参数": "details",
    "参数": "details",
    "功能介绍": "features",
    "功能": "features",
    "category": "category",
    "series": "series",
    "name": "name",
    "model": "model",
    "price": "price",
    "image": "productImage",
    "productimage": "productImage",
    "details": "details",
    "features": "features",
}


def build_import_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "产品导入"
    headers = [label for _, label in IMPORT_COLUMNS]
    ws.append(headers)
    ws.append([
        "前置过滤器",
        "清滤宝",
        "示例产品",
        "MODEL-001",
        1999,
        "https://example.com/product.png",
        "净水流量：3m³/h\n过滤精度：60微米",
        "过滤大颗粒杂质\n保护后端设备",
    ])
    ws.freeze_panes = "A2"
    widths = [16, 16, 24, 18, 12, 34, 36, 36]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0C6A58")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws["F1"].comment = Comment("可填图片 URL、本地图片路径，也可直接把图片插入到本列对应产品行。", "Codex")
    ws["G1"].comment = Comment("可换行填写。", "Codex")
    ws["H1"].comment = Comment("可换行填写。", "Codex")
    ws.row_dimensions[2].height = 58

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def read_current_products_data(username: str) -> dict:
    try:
        payload = json.loads(read_user_products(username))
    except json.JSONDecodeError:
        payload = default_custom_products()
    return normalize_custom_products(payload)


def import_products_from_payload(payload: dict, username: str, current_data: dict) -> dict:
    from openpyxl import load_workbook

    data_url = str(payload.get("dataUrl", ""))
    filename = sanitize_filename(str(payload.get("filename", "products.xlsx")))
    workbook_bytes = decode_data_url(data_url, 25 * 1024 * 1024)
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    ws = wb.active
    header_row, header_map = find_import_headers(ws)
    if not header_map:
        raise ValueError("没有找到导入表头")

    image_col = header_map.get("productImage")
    embedded_images = embedded_images_by_row(ws, image_col, header_row)
    base_products = load_base_products()
    base_by_id, base_matches = base_product_indexes(base_products)

    data = normalize_custom_products(current_data)
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    touched_categories: list[str] = []
    touched_ids: list[str] = []

    for row_index in range(header_row + 1, ws.max_row + 1):
        raw = read_import_row(ws, row_index, header_map)
        if not any(str(value or "").strip() for value in raw.values()) and row_index not in embedded_images:
            continue
        try:
            product = product_from_import_row(raw, username, row_index)
            if row_index in embedded_images:
                product["productImage"] = save_embedded_workbook_image(embedded_images[row_index], username, row_index)
            if not (product.get("name") or product.get("model")):
                skipped += 1
                errors.append(f"第 {row_index} 行缺少品名或型号")
                continue

            base_match = match_base_product(product, base_matches)
            addition_index = match_addition(product, data["additions"])
            if base_match:
                merged = {**base_match, **product, "id": base_match["id"]}
                data["overrides"][base_match["id"]] = clean_product(merged, base_match["id"])
                data["deletedProductIds"] = [item for item in data.get("deletedProductIds", []) if item != base_match["id"]]
                touched_ids.append(base_match["id"])
                updated += 1
            elif addition_index is not None:
                existing = data["additions"][addition_index]
                merged = {**existing, **product, "id": existing["id"]}
                data["additions"][addition_index] = clean_product(merged, existing["id"])
                touched_ids.append(existing["id"])
                updated += 1
            else:
                product_id = f"import-{int(time.time())}-{row_index}-{uuid.uuid4().hex[:8]}"
                data["additions"].append(clean_product({**product, "id": product_id}, product_id))
                touched_ids.append(product_id)
                created += 1
            if product.get("category"):
                touched_categories.append(product["category"])
        except Exception as exc:
            skipped += 1
            errors.append(f"第 {row_index} 行导入失败：{exc}")

    data["categoryOrder"] = append_unique(data["categoryOrder"], touched_categories)
    data["productOrder"] = append_unique(data["productOrder"], touched_ids)
    return {
        "data": normalize_custom_products(data),
        "imported": created + updated,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
    }


def decode_data_url(data_url: str, max_size: int) -> bytes:
    match = re.match(r"^data:[^;]+;base64,(.+)$", data_url, re.DOTALL)
    if not match:
        raise ValueError("文件格式不正确")
    content = base64.b64decode(match.group(1), validate=True)
    if not content or len(content) > max_size:
        raise ValueError("文件过大")
    return content


def find_import_headers(ws) -> tuple[int, dict[str, int]]:
    for row_index in range(1, min(ws.max_row, 10) + 1):
        header_map: dict[str, int] = {}
        for col_index in range(1, ws.max_column + 1):
            value = normalize_header(ws.cell(row_index, col_index).value)
            field = HEADER_ALIASES.get(value)
            if field and field not in header_map:
                header_map[field] = col_index
        if {"category", "name"}.issubset(header_map.keys()) or {"category", "model"}.issubset(header_map.keys()):
            return row_index, header_map
    return 0, {}


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def read_import_row(ws, row_index: int, header_map: dict[str, int]) -> dict:
    row = {}
    for field, col_index in header_map.items():
        value = ws.cell(row_index, col_index).value
        row[field] = "" if value is None else str(value).strip()
    return row


def product_from_import_row(row: dict, username: str, row_index: int) -> dict:
    category = row.get("category", "").strip() or "自定义产品"
    series = row.get("series", "").strip() or category
    product_image = normalize_import_image(row.get("productImage", ""), username, row_index)
    return {
        "category": category,
        "series": series,
        "name": row.get("name", "").strip(),
        "model": row.get("model", "").strip(),
        "price": normalize_price(row.get("price", "")),
        "productImage": product_image,
        "details": row.get("details", "").strip(),
        "features": row.get("features", "").strip(),
        "sourceSheet": "快速导入",
        "sourceRow": row_index,
    }


def embedded_images_by_row(ws, image_col: int | None, header_row: int) -> dict[int, object]:
    result = {}
    for image in getattr(ws, "_images", []):
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        if not marker:
            continue
        row = int(marker.row) + 1
        col = int(marker.col) + 1
        if row <= header_row:
            continue
        if image_col and col != image_col:
            continue
        result[row] = image
    return result


def save_embedded_workbook_image(image, username: str, row_index: int) -> str:
    image_bytes = image._data()
    fmt = str(getattr(image, "format", "") or "png").lower()
    ext = ".jpg" if fmt in ("jpeg", "jpg") else f".{fmt}"
    if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        ext = ".png"
    output_name = f"{user_file_stem(username)}-import-row{row_index}-{uuid.uuid4().hex[:8]}{ext}"
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    output_path = UPLOAD_DIR / output_name
    output_path.write_bytes(image_bytes)
    return f"assets/custom-products/{output_name}"


def normalize_import_image(value: str, username: str, row_index: int) -> str:
    image = str(value or "").strip()
    if not image:
        return ""
    if image.startswith(("http://", "https://", "assets/")):
        return image
    if image.startswith("/assets/"):
        return image.lstrip("/")
    if image.startswith("data:image/"):
        return save_data_url_image(image, username, f"import-row{row_index}")

    source = path_from_image_text(image)
    if source and source.exists() and source.is_file():
        ext = source.suffix.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
            return image
        output_name = f"{user_file_stem(username)}-import-row{row_index}-{uuid.uuid4().hex[:8]}{ext}"
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        output_path = UPLOAD_DIR / output_name
        shutil.copyfile(source, output_path)
        return f"assets/custom-products/{output_name}"
    return image


def path_from_image_text(value: str) -> Path | None:
    text = value.strip().strip('"')
    if text.startswith("file://"):
        parsed = urlparse(text)
        return Path(unquote(parsed.path).lstrip("/"))
    candidate = Path(text)
    if candidate.is_absolute():
        return candidate
    return ROOT / candidate


def save_data_url_image(data_url: str, username: str, stem: str) -> str:
    match = re.match(r"^data:(image/(?:png|jpeg|webp|gif));base64,(.+)$", data_url, re.DOTALL)
    if not match:
        return data_url
    mime_type, encoded = match.groups()
    image_bytes = base64.b64decode(encoded, validate=True)
    ext = {"image/png": ".png", "image/jpeg": ".jpg", "image/webp": ".webp", "image/gif": ".gif"}[mime_type]
    output_name = f"{user_file_stem(username)}-{stem}-{uuid.uuid4().hex[:8]}{ext}"
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    output_path = UPLOAD_DIR / output_name
    output_path.write_bytes(image_bytes)
    return f"assets/custom-products/{output_name}"


def load_base_products() -> list[dict]:
    if not BASE_CATALOG_FILE.exists():
        return []
    try:
        payload = json.loads(BASE_CATALOG_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    products = payload.get("products", [])
    return products if isinstance(products, list) else []


def base_product_indexes(products: list[dict]) -> tuple[dict[str, dict], list[dict]]:
    by_id = {}
    matches = []
    for product in products:
        if not isinstance(product, dict):
            continue
        clean = clean_product(product, str(product.get("id", "")))
        if clean.get("id"):
            by_id[clean["id"]] = clean
            matches.append(clean)
    return by_id, matches


def match_base_product(product: dict, base_products: list[dict]) -> dict | None:
    model = normalize_match_text(product.get("model", ""))
    name = normalize_match_text(product.get("name", ""))
    category = normalize_match_text(product.get("category", ""))
    if model:
        same_model = [item for item in base_products if normalize_match_text(item.get("model", "")) == model]
        if len(same_model) == 1:
            return same_model[0]
        for item in same_model:
            if normalize_match_text(item.get("category", "")) == category:
                return item
    if name and category:
        for item in base_products:
            if normalize_match_text(item.get("name", "")) == name and normalize_match_text(item.get("category", "")) == category:
                return item
    return None


def match_addition(product: dict, additions: list[dict]) -> int | None:
    model = normalize_match_text(product.get("model", ""))
    name = normalize_match_text(product.get("name", ""))
    category = normalize_match_text(product.get("category", ""))
    for index, item in enumerate(additions):
        if model and normalize_match_text(item.get("model", "")) == model and normalize_match_text(item.get("category", "")) == category:
            return index
        if name and normalize_match_text(item.get("name", "")) == name and normalize_match_text(item.get("category", "")) == category:
                return index
    return None


def update_user_product_price(username: str, product_id: str, price: int | float) -> dict:
    data = normalize_custom_products(json.loads(read_user_products(username)))
    for index, product in enumerate(data["additions"]):
        if product.get("id") == product_id:
            updated = clean_product({**product, "price": price}, product_id)
            data["additions"][index] = updated
            write_user_products(username, data)
            return updated

    base_by_id, _ = base_product_indexes(load_base_products())
    base_product = base_by_id.get(product_id)
    override = data["overrides"].get(product_id)
    if not base_product and not override:
        raise ValueError("product not found")

    updated = clean_product({**(base_product or {}), **(override or {}), "id": product_id, "price": price}, product_id)
    data["overrides"][product_id] = updated
    write_user_products(username, data)
    return updated


def normalize_match_text(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def append_unique(existing: list, values: list[str]) -> list[str]:
    result = clean_order(existing)
    for value in values:
        item = str(value or "").strip()
        if item and item not in result:
            result.append(item)
    return result


def render_pdf(html: str, width: int, height: int) -> bytes:
    chrome = find_chrome()
    if not chrome:
        raise RuntimeError("Chrome or Edge was not found")

    with tempfile.TemporaryDirectory(prefix="quote-pdf-") as tmp:
        tmp_path = Path(tmp)
        html_path = tmp_path / "quote-export.html"
        pdf_path = tmp_path / "quote-export.pdf"
        user_data = tmp_path / "chrome-profile"
        html_path.write_text(html, encoding="utf-8")

        subprocess.run(
            [
                chrome,
                "--headless=new",
                "--disable-gpu",
                "--no-sandbox",
                "--no-first-run",
                "--disable-extensions",
                f"--user-data-dir={user_data}",
                f"--window-size={width},{height}",
                f"--print-to-pdf={pdf_path}",
                html_path.as_uri(),
            ],
            cwd=str(ROOT),
            check=True,
            timeout=60,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

        if not pdf_path.exists():
            raise RuntimeError("Chrome did not create a PDF")
        return pdf_path.read_bytes()


def sanitize_filename(name: str) -> str:
    keep = "".join("_" if char in '\\/:*?"<>|' else char for char in name).strip()
    if not keep:
        keep = "报价单.pdf"
    if not keep.lower().endswith(".pdf"):
        keep += ".pdf"
    return keep


def quote_filename(name: str) -> str:
    return "".join(f"%{byte:02X}" for byte in name.encode("utf-8"))


def cleanup_exports() -> None:
    cutoff = time.time() - 24 * 60 * 60
    for path in EXPORT_DIR.glob("*.pdf"):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            pass


def normalize_custom_products(payload: object) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("payload must be an object")

    overrides = payload.get("overrides", {})
    additions = payload.get("additions", [])
    category_order = payload.get("categoryOrder", [])
    product_order = payload.get("productOrder", [])
    deleted_product_ids = payload.get("deletedProductIds", [])
    if not isinstance(overrides, dict):
        overrides = {}
    if not isinstance(additions, list):
        additions = []
    if not isinstance(category_order, list):
        category_order = []
    if not isinstance(product_order, list):
        product_order = []
    if not isinstance(deleted_product_ids, list):
        deleted_product_ids = []

    clean_overrides = {}
    for product_id, product in overrides.items():
        if isinstance(product_id, str) and isinstance(product, dict):
            clean_overrides[product_id] = clean_product(product, product_id)

    clean_additions = []
    for product in additions:
        if isinstance(product, dict):
            clean = clean_product(product)
            if clean.get("id"):
                clean_additions.append(clean)

    return {
        "overrides": clean_overrides,
        "additions": clean_additions,
        "categoryOrder": clean_order(category_order),
        "productOrder": clean_order(product_order),
        "deletedProductIds": clean_order(deleted_product_ids),
    }


def clean_order(values: list) -> list[str]:
    clean = []
    seen = set()
    for value in values:
        if not isinstance(value, str):
            continue
        item = value.strip()
        if item and item not in seen:
            seen.add(item)
            clean.append(item)
    return clean


def clean_product(product: dict, fallback_id: str = "") -> dict:
    fields = [
        "id",
        "category",
        "series",
        "name",
        "model",
        "price",
        "priceLabel",
        "productImage",
        "installImage",
        "featureImage",
        "details",
        "installText",
        "features",
        "sourceSheet",
        "sourceRow",
    ]
    clean = {}
    for field in fields:
        value = product.get(field)
        if field == "price":
            clean[field] = normalize_price(value)
        elif field == "sourceRow":
            try:
                clean[field] = int(value)
            except (TypeError, ValueError):
                clean[field] = 0
        else:
            clean[field] = "" if value is None else str(value)
    if not clean["id"]:
        clean["id"] = fallback_id
    return clean


def normalize_quote_state(payload: object) -> dict:
    if not isinstance(payload, dict):
        return default_quote_state()
    quote = payload.get("quote", [])
    customer = payload.get("customer", {})
    terms = payload.get("terms", {})
    offer = payload.get("offer", "")
    quote_code = payload.get("quoteCode", "")
    owner = normalize_username(str(payload.get("owner", "")))
    updated_by = normalize_username(str(payload.get("updatedBy", "")))
    updated_at = normalize_timestamp(payload.get("updatedAt"))

    if not isinstance(quote, list):
        quote = []
    if not isinstance(customer, dict):
        customer = {}
    if not isinstance(terms, dict):
        terms = {}

    state = {
        "quote": quote[:200],
        "customer": {str(key): "" if value is None else str(value) for key, value in customer.items()},
        "offer": "" if offer is None else str(offer),
        "terms": {str(key): "" if value is None else str(value) for key, value in terms.items()},
        "quoteCode": "" if quote_code is None else str(quote_code),
    }
    if owner:
        state["owner"] = owner
    if updated_by:
        state["updatedBy"] = updated_by
    if updated_at:
        state["updatedAt"] = updated_at
    return state


def normalize_price(value: object) -> int | float | None:
    if value in ("", None):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else number


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Serve the quote website with PDF export.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=4173)
    args = parser.parse_args()

    mimetypes.add_type("application/javascript", ".js")
    mimetypes.add_type("image/webp", ".webp")
    ensure_users_file()
    address = (args.host, args.port)
    httpd = ThreadingHTTPServer(address, QuoteHandler)
    print(f"Serving {ROOT} at http://{args.host}:{args.port}/")
    httpd.serve_forever()
